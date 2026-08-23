#!/usr/bin/env python3
"""Turn scenario-test.mjs output into a workbook Neha can read and grade.

    python3 scripts/results-to-xlsx.py baseline.json [after.json] -o reflect-respond-test.xlsx

One JSON = a single run. Two = before/after a prompt change, and the Reflect sheet
gains a `Run` column so the same scenario's old and new lines sit together.

The two auto-flag columns (Repeat / Echo) are cheap heuristics, there to point the
eye at the rows worth reading, never to replace the read.
"""
import json
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
FLAG_FILL = PatternFill("solid", fgColor="FCE4E4")   # auto-flagged row
FILL_ME = PatternFill("solid", fgColor="FFFF00")     # columns for her to fill in
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STOP = set(
    "a an and are as at be been but by can could did do does for from get got had has have "
    "he her hers him his how i if in is it its just me might my no not of on or our out she "
    "so some that the their them then there they this to too up us was we were what when "
    "which who why will with would you your yours am been being does doing".split()
)

# The mirror openers: a line starting this way is usually about to hand her own
# words back. Drawn from the baseline run, not invented. NOTE this is a metric in
# this script only; the flow's old "echo" beat (say her words back, then "did I
# get that right?") was removed and is unrelated.
ECHO_OPENERS = re.compile(
    r"^(it (is|was|feels|makes sense|can be|could be that you|might be that you|is hard|is a lot)"
    r"|you (might|may|could|are|feel|must)\b"
    r"|maybe you (feel|are)\b"
    r"|perhaps you (feel|are)\b"
    r"|this (is|might be a feeling|could be a feeling)"
    r"|that (is|sounds)\b)",
    re.I,
)


def words(s):
    return [w for w in re.findall(r"[a-z']+", s.lower()) if w not in STOP]


def jaccard(a, b):
    sa, sb = set(a), set(b)
    return len(sa & sb) / len(sa | sb) if sa | sb else 0.0


def echo_score(line, her_text):
    """How much of the line is her own vocabulary handed back."""
    lw = words(line)
    if not lw:
        return 0.0
    hw = set(words(her_text))
    return sum(1 for w in lw if w in hw) / len(lw)


def load(path):
    with open(path) as f:
        return json.load(f)


def flag_reflect(rows):
    """Add Repeat / Echo flags. Repeat compares each line to every earlier line of
    the same scenario in the same run, which is what she actually sees stacked up."""
    seen = {}
    for r in rows:
        key = (r.get("run", ""), r["scenario"])
        prior = seen.setdefault(key, [])
        line = r["line"]
        toks = words(line)
        dup = ""
        for p_line, p_toks, p_card, p_pull in prior:
            if line.strip().lower() == p_line.strip().lower():
                dup = f"SAME as {p_card} pull {p_pull}"
                break
            if len(toks) >= 3 and jaccard(toks, p_toks) >= 0.55:
                dup = f"near-dup of {p_card} pull {p_pull}"
                break
        r["repeat"] = dup
        es = echo_score(line, r["text"])
        r["echo"] = (
            "opener + her words"
            if ECHO_OPENERS.match(line.strip()) and es >= 0.30
            else "mirror opener"
            if ECHO_OPENERS.match(line.strip())
            else "her own words back"
            if es >= 0.45
            else ""
        )
        prior.append((line, toks, r["card"], r["pull"]))
    return rows


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_sheet(ws, headers, rows, widths, fill_from=None, flag_col=None):
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header(ws, len(headers))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        flagged = flag_col is not None and bool(row[flag_col].value)
        for i, cell in enumerate(row):
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if fill_from is not None and i >= fill_from:
                cell.fill = FILL_ME
            elif flagged:
                cell.fill = FLAG_FILL


def main():
    argv = sys.argv[1:]
    out = "reflect-respond-test.xlsx"
    if "-o" in argv:
        i = argv.index("-o")
        out = argv[i + 1]
        argv = argv[:i] + argv[i + 2:]
    args = [a for a in argv if not a.startswith("-")]
    runs = []
    for p in args:
        d = load(p)
        for r in d["reflect"]:
            r["run"] = d["tag"]
        for r in d["respond"]:
            r["run"] = d["tag"]
        runs.append(d)
    two = len(runs) > 1

    reflect = flag_reflect([r for d in runs for r in d["reflect"]])
    respond = [r for d in runs for r in d["respond"]]

    wb = Workbook()

    # --- Read me ---------------------------------------------------------
    ws = wb.active
    ws.title = "Read me"
    lines = [
        ("Reflect + Respond prompt test", True),
        ("", False),
        (f"Run(s): {', '.join(d['tag'] for d in runs)}   ·   model: {runs[0]['model']}   ·   {runs[0]['at'][:16]}", False),
        ("", False),
        ("What this is", True),
        ("Every line here came out of the real prompts in src/lib/moment-prompts.ts, through the same", False),
        ("model and backend the app uses. Nothing is mocked. Regenerate with:", False),
        ("    node --experimental-strip-types scripts/scenario-test.mjs --out run.json --tag mytag", False),
        ("    python3 scripts/results-to-xlsx.py run.json -o out.xlsx", False),
        ("", False),
        ("Sheets", True),
        ("Reflect   every read she would see, one per row, three pulls deep (pull 2 and 3 = 'Show me more')", False),
        ("Respond   the draft she gets after choosing a move", False),
        ("Expand    what one read looks like opened out. The prompt exists, the tap does not yet", False),
        ("Summary   counts, live formulas over the Reflect sheet", False),
        ("Findings  what the run showed and what was done about it. Read this first", False),
        ("", False),
        ("Two runs are stacked in one sheet. Filter the Run column:", False),
        ("  baseline  the prompts as they were on 19 Aug, before any change", False),
        ("  fixed     the prompts now, with the routing and thinking-budget changes", False),
        ("'Read as' is what detectTimeframe() made of her words. Blank on baseline, which predates it.", False),
        ("", False),
        ("The two auto-flag columns", True),
        ("Repeat  this line is a near-duplicate of an earlier line in the same scenario. Pink = flagged.", False),
        ("Adds nothing new   this line gives her back her own words instead of adding something.", False),
        ("                   A measurement label for this sheet only. NOT the old Echo beat, which is gone from the flow.", False),
        ("Both are pointers, not verdicts. A flagged line can be fine; an unflagged line can be filler.", False),
        ("", False),
        ("Grading (the yellow columns are yours)", True),
        ("Keep?          y / n / maybe", False),
        ("What's missing free text: what you wanted this line to say instead", False),
        ("Example row:   Keep? = n    What's missing = 'she has heard this for 15 years, it is not news'", False),
        ("", False),
        ("Filter the Reflect sheet by 'When' to compare acute against long-standing.", False),
    ]
    for i, (t, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=t)
        c.font = Font(name=FONT, size=11, bold=bold)
    ws.column_dimensions["A"].width = 110

    # --- Reflect ---------------------------------------------------------
    ws = wb.create_sheet("Reflect")
    heads = (["Run"] if two else []) + [
        "Scenario", "When", "Read as", "Her words", "Feeling", "Cards routed", "Card", "Card asks",
        "Pull", "Line #", "The line she reads", "Words", "Repeat", "Adds nothing new", "Keep?", "What's missing",
    ]
    rows = [
        ([r["run"]] if two else [])
        + [
            r["scenario"], r["when"], r.get("read") or "n/a (pre-fix run)", r["text"], r["feeling"], r["routed"],
            r["card"], r["title"], r["pull"], r["lineIdx"], r["line"], r["words"], r["repeat"],
            r["echo"], "", "",
        ]
        for r in reflect
    ]
    widths = ([8] if two else []) + [20, 18, 12, 46, 12, 34, 14, 26, 6, 6, 54, 7, 20, 18, 8, 30]
    write_sheet(ws, heads, rows, widths, fill_from=len(heads) - 2, flag_col=len(heads) - 4)

    # --- Respond ---------------------------------------------------------
    ws = wb.create_sheet("Respond")
    heads = (["Run"] if two else []) + [
        "Scenario", "When", "Her words", "Feeling", "Her move", "The draft", "Words", "Keep?", "What's missing",
    ]
    rows = [
        ([r["run"]] if two else [])
        + [r["scenario"], r["when"], r["text"], r["feeling"], r["act"], r["draft"], r["words"], "", ""]
        for r in respond
    ]
    widths = ([8] if two else []) + [20, 18, 46, 12, 20, 60, 7, 8, 30]
    write_sheet(ws, heads, rows, widths, fill_from=len(heads) - 2)

    # --- Expand (tap a read to open it) ----------------------------------
    expand = [r for d in runs for r in d.get("expand", []) if (r.setdefault("run", d["tag"]) or True)]
    if expand:
        ws = wb.create_sheet("Expand")
        heads_e = (["Run"] if two else []) + [
            "Scenario", "When", "Her words", "Card", "The read she tapped",
            "Opened out", "Words", "Keep?", "What's missing",
        ]
        rows_e = [
            ([r["run"]] if two else [])
            + [r["scenario"], r["when"], r["text"], r["card"], r["read"], r["expanded"], r["words"], "", ""]
            for r in expand
        ]
        write_sheet(
            ws, heads_e, rows_e,
            ([8] if two else []) + [20, 18, 40, 14, 44, 70, 7, 8, 30],
            fill_from=len(heads_e) - 2,
        )

    # --- Summary (live formulas over Reflect) ----------------------------
    ws = wb.create_sheet("Summary")
    n = len(reflect) + 1
    rcol = {}
    rheads = (["Run"] if two else []) + [
        "Scenario", "When", "Read as", "Her words", "Feeling", "Cards routed", "Card", "Card asks",
        "Pull", "Line #", "The line she reads", "Words", "Repeat", "Adds nothing new", "Keep?", "What's missing",
    ]
    for i, h in enumerate(rheads, start=1):
        rcol[h] = get_column_letter(i)
    rep, ech, wrd, card = rcol["Repeat"], rcol["Adds nothing new"], rcol["Words"], rcol["Card"]
    ws.append(["Measure", "Value", "What it means"])
    body = [
        ("Lines generated", f"=COUNTA(Reflect!{rcol['The line she reads']}2:{rcol['The line she reads']}{n})", "one row per read she would see"),
        ("Flagged as a repeat", f"=COUNTIF(Reflect!{rep}2:{rep}{n},\"<>\")", "near-duplicate of an earlier line in the same scenario"),
        ("Repeat rate", f"=IFERROR(B3/B2,0)", "share of reads that are recycled"),
        ("Flagged as adding nothing", f"=COUNTIF(Reflect!{ech}2:{ech}{n},\"<>\")", "gives her back her own words instead of adding"),
        ("Adds-nothing rate", f"=IFERROR(B5/B2,0)", "share of reads that add nothing new"),
        ("Empty / declined", f"=COUNTIF(Reflect!{rcol['The line she reads']}2:{rcol['The line she reads']}{n},\"(EMPTY*\")", "the card came up blank for her"),
        ("Average words per read", f"=IFERROR(ROUND(AVERAGE(Reflect!{wrd}2:{wrd}{n}),1),0)", "how much she gets to read at once"),
        ("Longest read", f"=MAX(Reflect!{wrd}2:{wrd}{n})", "the ceiling on depth"),
        ("Reads on the pattern card", f"=COUNTIF(Reflect!{card}2:{card}{n},\"pattern\")", "the card that fires first on a long-standing issue"),
    ]
    for r in body:
        ws.append(list(r))
    style_header(ws, 3)
    for row in ws.iter_rows(min_row=2, max_col=3):
        for cell in row:
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in (4, 6):
        ws.cell(row=r, column=2).number_format = "0.0%"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60

    # --- Findings --------------------------------------------------------
    ws = wb.create_sheet("Findings")
    ws.append(["#", "What is wrong", "Evidence in this run", "Why it matters to her", "Fix"])
    for r in FINDINGS:
        ws.append(list(r))
    style_header(ws, 5)
    for i, w in enumerate([5, 34, 50, 44, 44], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_col=5):
        for cell in row:
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    wb.save(out)
    print(f"wrote {out}: {len(reflect)} reflect rows, {len(respond)} respond rows")


FINDINGS = [
    (1, "The pattern card never reached her",
     "Every pattern row in the baseline run starts with the word 'none' followed by a wall of lines. moment-ai.ts reads a leading 'none' as a decline and throws the whole reply away.",
     "pattern is the FIRST card on any long-standing issue. The moment she brought something she had carried for years, the one card built for it rendered authored fallback copy.",
     "FIXED. The prompt asked for 'themes from her past entries' the app never sent. moment.tsx now sends her same-thread history (themesClause, capped at 4 entries) and the prompt asks for the one thing those times have in common."),
    (2, "Nothing in the flow knew when it happened",
     "Baseline: longstanding_mum (15 years) routed to 'simpler', which offered 'Your mum might be worried about you'. acute_partner (mid-fight, shaking) got three analysis cards before anything steadying.",
     "A 15-year pattern does not need a gentler reason for one incident, she has run that a thousand times. A woman shaking mid-fight cannot take in a lens.",
     "FIXED. detectTimeframe() reads acute / longstanding / recent. Long-standing leads with pattern, rule, need, whose_weight and demotes simpler; acute leads with need and signal. TIME_NOTE goes into every reflect and respond prompt. Reorder only, so nothing is dropped."),
    (3, "'Show me more' handed back the same lines",
     "Baseline vague_mood / need pull 2 repeated all three of pull 1 word for word. The rule card returned byte-identical output on all three pulls.",
     "That is the exact moment she stops trusting it.",
     "FIXED, 9% to 4%. The do-not-repeat rule moved into the instruction, and the rule card now feeds its last chain back into its own re-roll (moment.tsx lastRule)."),
    (4, "Cards came back empty",
     "3 empty cards in the baseline, 7 after the first prompt pass.",
     "She taps a card and nothing arrives.",
     "FIXED, now 0. Cause was thinkingBudget 0 with a 256-token cap. Raising thinking to 512 with a matching cap removed every empty. Costs about 3.1s a call, and the flow prefetches."),
    (5, "Third person leaked into the reads",
     "Baseline acute_short / need: 'to know she is important to him', 'to know she can count on him'.",
     "It reads as if the app is discussing her with somebody else.",
     "MOSTLY FIXED. REFLECT_SAFETY now says the turn describes her in the third person but the reply goes straight to her. Remaining hits are legitimate ('she' meaning her mum or her friend)."),
    (6, "Reads give her back her own words (this sheet's 'Adds nothing new' column, not the removed Echo beat)",
     "39% of baseline lines. 37% after both prompt passes. Concentrated: signal 71%, middle 40%. signal now restates the EVENT instead of the feeling ('You were mid-sentence, and he chose to walk away').",
     "She wrote those words. A read that repeats them costs her a tap and gives her nothing, and it is the main reason the flow reads as boring.",
     "NOT FIXED, and two prompt rewrites did not move it. Ban a shape and the model finds another. Worth trying next: split signal into two calls (name what is at stake, then write from it), or cut the card. Do not spend another pass on wording."),
    (7, "She cannot read more than one line",
     "Baseline reads average 13.9 words and the only 'more' on offer is more one-liners.",
     "When a read finally lands there is nowhere to go with it. Depth is what she came for.",
     "PROMPT BUILT, UI NOT WIRED. reflect_expand takes the read she tapped and opens that one thought out in 3 or 4 sentences. See the Expand sheet. Needs a tap target on a read in moment.tsx."),
    (8, "Respond drafts run to one template, and paste the feeling chip",
     "Three of nine baseline drafts are 'When you X, it made me feel Y, I need you to Z'. longstanding_mum: 'I feel unappreciated when you comment on my weight'. Unappreciated came from the chip, not from her.",
     "A template is not her voice, and the chip puts a word in her mouth in a message she may actually send.",
     "FIXED in the prompt. act_help now says the feeling is context for tone only, and to let the shape follow the move. Read the Respond sheet to judge it."),
    (9, "The Gemini key in the app bundle answers without App Check",
     "This harness calls the app's own Vertex endpoint using only the API_KEY out of GoogleService-Info.plist, from a laptop, with no attestation.",
     "Anyone who unpacks the IPA can spend the project's Gemini quota. Not a Reflect problem, found while building the harness.",
     "OPEN. This is launch-prep gate #1. Either enforce App Check on the Vertex AI API or proxy the key."),
]


if __name__ == "__main__":
    main()
